import spacy
import jsonlines
import openpyxl
import os
import string
from nltk.tokenize import word_tokenize
import actions


def SimilarityFinder(text):
    
    
    all_stopwords = actions.nlp.Defaults.stop_words

    message = text
    answer = dict()
    choices = dict()
    answerkey = dict()
    confd_score = dict()

        

    with jsonlines.open('dataset/sd.jsonl') as reader:
        print("entered in json")
        text_tokens = word_tokenize(message)
        tokens_without_sw= [word for word in text_tokens if not word in all_stopwords]
        message_sw = " ".join(tokens_without_sw)

        for obj in reader:
            # This will remove stopwords from the user input.
            
            test_score = actions.nlp(obj['question']).similarity(actions.nlp(message_sw)) 

            if(test_score > 0.85):
                answer[obj['id']] = obj['answer']
                answerkey[obj['id']]= obj['answerkey']
                confd_score[obj['id']] = test_score
        print("exit from json")       
   
    wb = openpyxl.load_workbook('score/score.xlsx');
    workbook = wb.active
        
    if( len(answerkey) == 0 ):
        fa =  "Oops! I could not find the answer." 
        print( "No Answer found for the question." )
        if(workbook.max_row > 1):
            data = (workbook.max_row , "NA" , text , "NA" , "NA" , "Below 85 %" )
        else:
            data = (1 , "NA" , text , "NA" , "NA" , "Below 85 %" )
    else:
        sort_ans = sorted(confd_score.items(), key = lambda x: x[1], reverse=True)
        id = sort_ans[0][0]
        answerkey_txt= answerkey[id]
        answer_txt = answer[id]
        confd_scoretxt = float(confd_score[id])
        
        print("Answer is :" +answer_txt + "AnswerKey :" +answerkey_txt ) 
        fa = answer_txt.capitalize()
        
        if(workbook.max_row > 1):
            data = (workbook.max_row , id , text , answerkey_txt , answer_txt , confd_scoretxt )
        else:
            data = (1 , id , text , answerkey_txt , answer_txt , confd_scoretxt )

    
    return [fa , data]

def writeToExcel(data):
    wb = openpyxl.load_workbook('score/score.xlsx');
    workbook = wb.active

    if(workbook.max_row > 1):
        workbook.append(data)            
    else:
        data1 = ("S.No." , "ID" , "Question" , "Correct Answer Key" , "Correct Answer" , "Confidence Score" )
        workbook.append(data1)
        workbook.append(data)
        
    wb.save('score/score.xlsx');
    


